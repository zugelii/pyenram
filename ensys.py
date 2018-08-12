#-*- coding=utf-8 -*-

import MySQLdb as mdb
import sys, xlrd
from openpyxl import load_workbook
import time, threading
from pygments.lexers import _tsql_builtins

#
#


conn = mdb.connect(host = "127.0.0.1",
                   user = 'root',
                   passwd = '123456',
                   db = 'enpdu')

cursor = conn.cursor()
cursor.execute('SELECT VERSION()')
data = cursor.fetchone()


db_lock = threading.Lock()

def db_insert_data(db, sql, val):           #执行SQL语句，插入数据
    try:
        cursor_l = db.cursor()
        cursor_l.execute(sql, val)
        db.commit()
    except:
        db.rollback()
        print("unable insert data")        

def open_excel(file = "pdulist.xlsx"):
    try:
        wb = load_workbook(file)
        return wb
    except:
        print("not pdulist.xlsx")
    
#wb = open_excel("pdulist.xlsx")    
def copy_excel_pdu_list(wb):
    sheetnames = wb.sheetnames;
    for sheetname in sheetnames:        
        sheet = wb.get_sheet_by_name(sheetname)
        if(sheet['A1'].value != "PDU IP"):
            continue
        for row in range(sheet.max_column - 2):   #从第二行开始         
            #for col in sheet[str(row + 2)]:  #获取一行的数据
            #sql = "insert into `pdu list`(`PDU IP`, `User Name`, Password, SKU, `Serial Num`, Tag)  values(%s,%s,%s,%s,%s,%s)" 
            sql = "insert into `pdu list`(`PDU IP`, `User Name`, `Password`, `SKU`, `Serial Num`, `Tag`)  values(%s, %s, %s, %s, %s, %s)"
            val = (sheet['A' + str(row + 2)].value, sheet['B' + str(row + 2)].value, sheet['C' + str(row + 2)].value, \
                   sheet['D' + str(row + 2)].value, sheet['E' + str(row + 2)].value, sheet['F' + str(row + 2)].value)     
            db_insert_data(conn, sql, val)
   
def db_get_data_pdu_database(sku):  #获取OID信息
    sql = "select * from `pdu list` where SKU=" + "\"" + sku + "\""   #需要转义字符
    try:
        global cursor
        cursor.execute(sql)
        result = cursor.fetchone()
        return result
    except:
        print("can not read pdu database")
     

def db_get_data_pdu_list(row, table):
    #sql = 'SELECT' + row + 'FROM' + table
    #sql = "select * from taskqueue"
    #sql = "select * from " + table + "where SKU="P24G01M"
    sku = "\"P24G01M\""
    sql = "select * from `pdu list` where SKU=" + sku   #需要转义字符
    try:
        global cursor
        cursor.execute(sql)
        results = cursor.fetchall()
        print("table has row:%d" % len(results))
        for rows_data in results:
            pdu_ip = rows_data[1]
            user_name = rows_data[2]
            passwd = rows_data[3]
            sku    = rows_data[4]
            print("ip:%s name:%s pw:%s sku:%s" %(pdu_ip, user_name, passwd, sku))
    except:
        print("unable to fetch data")       



def db_insert_data1():
    #sql = 'INSERT INTO' + table + 'VALUES(%s,%s,%d,%d)'
    #sql = "insert into `test user`(user, sex, `id d`)  values(%s,%s,%s)"
    #val1 = ("test8","2232",45)
    sql = "insert into user(`User Name`, `Card Number`, `Pin Number`, `Status`)  values('%s','%s','%d','%d')" %("test2","2232",134,0)
    try:
        #cursor.execute(sql, value)
        cursor.execute(sql)
        conn.commit()
    except:
        conn.rollback()
        print("unable insert data")


def db_delete_data(table, row):
    sql = "delete from " + table + " where id = " + row
    #sql = "delete from testuser where id = 9"
    try:
        cursor.execute(sql)
        conn.commit()
    except:
        conn.rollback()
        print("cant delete row")
    
    
n = 0
def check_pdu_sensor():
    print("thread %s is running" % threading.current_thread().name)
    db_lock.acquire()
    try:
        global n
        while n < 5:
            n = n + 1
            print("thread %s  >>> n:%d" %(threading.current_thread().name, n))
            time.sleep(2)
    finally:
        db_lock.release()   #释放锁
    print("thread %s end" % threading.current_thread().name)

def check_pdu_sensor2():
    print("thread %s is running" % threading.current_thread().name)
    db_lock.acquire()
    try:
        global n
        while n < 5:
            n = n + 1
            print("thread %s  >>> n:%d" %(threading.current_thread().name, n))
            time.sleep(2)
    finally:
        db_lock.release()   #释放锁
        
    print("thread %s end" % threading.current_thread().name)


def start_check_sensor_thread():
    t = threading.Thread(target=check_pdu_sensor, name = "pdu sensor thread")
    t.start()
   
    t2 = threading.Thread(target=check_pdu_sensor2, name = "pdu sensor thread2")
    t2.start()
    t.join()
    t2.join()    
    
def main():
    print ("Database version : %s " % data)
    conn.close()
    
    
if __name__ == '__main__':
    #db_get_data('`Card Number`','user')
    #db_insert_data(('test','222332',1234,1),'user')
    #db_delete_data("testuser", "10")
    #wb = open_excel("pdulist.xlsx")
    #copy_excel_pdu_list(wb)
    #db_insert_data1()
    db_get_data_pdu_list(1, "`pdu list`")
    main()



